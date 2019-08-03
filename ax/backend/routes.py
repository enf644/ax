"""All routes of ax-core and ax-admin"""

import os
import sys
import asyncio
import uuid

from sanic import response
from loguru import logger
import aiopubsub
from graphql_ws.websockets_lib import WsLibSubscriptionServer
from sanic_graphql import GraphQLView
from graphql.execution.executors.asyncio import AsyncioExecutor
import ujson as json

import backend.cache as ax_cache
import backend.schema as ax_schema
import backend.misc as ax_misc
import backend.pubsub as ax_pubsub
from backend.tus import tus_bp
import backend.model as ax_model
from backend.model import AxForm, AxField
import backend.schemas.form_schema as form_schema
import backend.scheduler as ax_scheduler
import backend.dialects as ax_dialects

this = sys.modules[__name__]

loop = asyncio.new_event_loop()
actions_loop = asyncio.new_event_loop()
app = None
graphql_view = None
dummy_view = None


class AxGraphQLView(GraphQLView):
    """ Extends GraphQLView to output GQL errors and schema updates"""

    def __init__(self, **kwargs):
        super().__init__(schema=ax_schema.schema,
                         graphiql=False,
                         enable_async=True,
                         executor=AsyncioExecutor(loop=this.loop))

    @staticmethod
    def format_error(error):
        logger.error(error)
        return GraphQLView.format_error(error)


def init_graphql_view():  # pylint: disable=unused-variable
    """Initiate graphql"""
    this.graphql_view = AxGraphQLView.as_view()
    this.app.add_route(this.graphql_view, '/api/graphql')



async def do_action():
    """ Test function """
    from backend.schema import schema

    values = {
        "searchString": "Chateau Lafite Rothschild 1996",
        "supplier": "9c198568469447268d62b395badeb71f",
        "loader": "0eb38c45e9954442995c6205f2e3f7eb"
    }

    query = f"""
        mutation(
            $formDbName: String,
            $actionDbName: String,
            $values: String
        ){{
            doAction(
                formDbName: $formDbName
                actionDbName: $actionDbName
                values: $values
            ) {{
                form {{
                guid
                dbName
                }}
                newGuid
                messages
                ok
            }}
        }}
    """
    variables = {
        "formDbName": "Stock",
        "actionDbName": "addDraft",
        "values": json.dumps(values)
    }
    result = schema.execute(query, variables=variables)
    test_str = json.dumps(result.data, sort_keys=True, indent=4)
    print(test_str)


def init_routes(sanic_app):
    """Innitiate all Ax routes"""
    try:
        # Add tus upload blueprint
        sanic_app.blueprint(tus_bp)
        # Add web-socket subscription server
        subscription_server = WsLibSubscriptionServer(ax_schema.schema)

        @sanic_app.route(
            '/api/file/<form_guid>/<row_guid>/<field_guid>/<file_name>', methods=['GET'])
        async def db_file_viewer(request, form_guid, row_guid, field_guid, file_name):  # pylint: disable=unused-variable
            """ Used to display files that are stored in database.
                Used in fields like AxImageCropDb"""
            del request, form_guid
            safe_row_guid = str(uuid.UUID(str(row_guid)))
            ax_field = ax_model.db_session.query(AxField).filter(
                AxField.guid == uuid.UUID(field_guid)
            ).first()
            field_value = await ax_dialects.dialect.select_field(
                form_db_name=ax_field.form.db_name,
                field_db_name=ax_field.db_name,
                row_guid=safe_row_guid)
            return response.raw(
                field_value, content_type='application/octet-stream')

        @sanic_app.route(
            '/api/file/<form_guid>/<row_guid>/<field_guid>/<file_guid>/<file_name>',
            methods=['GET'])
        async def file_viewer(  # pylint: disable=unused-variable
                request, form_guid, row_guid, field_guid, file_guid, file_name):
            """ Used to display files uploaded and stored on disk.
                Displays temp files too. Used in all fields with upload"""
            del request

            # if row_guid is null -> display from /tmp without permissions
            if not row_guid or row_guid == 'null':
                tmp_dir = os.path.join(ax_misc.tmp_root_dir, file_guid)
                file_name = os.listdir(tmp_dir)[0]
                temp_path = os.path.join(tmp_dir, file_name)
                return await response.file(temp_path)

            # get AxForm with row values
            ax_form = ax_model.db_session.query(AxForm).filter(
                AxForm.guid == uuid.UUID(form_guid)
            ).first()
            ax_form = await form_schema.set_form_values(
                ax_form=ax_form, row_guid=row_guid)

            # Get values from row, field
            field_values = None
            for field in ax_form.fields:
                if field.guid == uuid.UUID(field_guid):
                    if field.value:
                        field_values = json.loads(field.value)

            # Find requested file in value
            the_file = None
            for file in field_values:
                if file['guid'] == file_guid:
                    the_file = file

            if not the_file:
                return response.text("", status=404)

            # if file exists -> return file
            row_guid_str = str(uuid.UUID(row_guid))
            file_path = os.path.join(
                ax_misc.uploads_root_dir,
                'form_row_field_file',
                form_guid,
                row_guid_str,
                field_guid,
                the_file['guid'],
                the_file['name'])
            if not os.path.lexists(file_path):
                return response.text("", status=404)
            return await response.file(file_path)


        @sanic_app.route('/<path:path>')
        def index(request, path):  # pylint: disable=unused-variable
            """ This is MAIN ROUTE. (except other routes listed in this module).
                All requests are directed to Vue single page app. After that Vue
                handles routing."""
            del request, path
            absolute_path = ax_misc.path('dist/index.html')
            return response.html(open(absolute_path).read())

        @sanic_app.route('/draw_ax')
        async def draw_ax(request):  # pylint: disable=unused-variable
            """ Outputs bundle.js. Used when Ax web-components
                are inputed somewhere. Users can use this url for <script> tag
                """
            del request
            absolute_path = ax_misc.path('dist/static/js/ax-bundle.js')
            return await response.file(
                absolute_path,
                headers={
                    'Content-Type': 'application/javascript; charset=utf-8'
                }
            )

        @sanic_app.websocket('/api/subscriptions', subprotocols=['graphql-ws'])
        async def subscriptions(request, web_socket):  # pylint: disable=unused-variable
            """Web socket route for graphql subscriptions"""
            del request
            try:
                # TODO: Why socket error exception occurs without internet
                await subscription_server.handle(web_socket)
                return web_socket
            except asyncio.CancelledError:
                pass
                # logger.exception('Socket error')

        @sanic_app.route('/api/test')
        async def test(request):  # pylint: disable=unused-variable
            """Test function"""
            del request
            this.test_schema = 'IT WORKS'
            ax_pubsub.publisher.publish(
                aiopubsub.Key('dummy_test'), this.test_schema)
            return response.text(this.test_schema)

        @sanic_app.route('/api/set')
        async def cache_set(request):  # pylint: disable=unused-variable
            """Cache Test function"""
            del request
            obj = ['one', 'two', 'three']
            await ax_cache.cache.set('user_list', obj)
            return response.text('Cache SET' + str(obj))

        @sanic_app.route('/api/get')
        async def cache_get(request):  # pylint: disable=unused-variable
            """Cache Test function"""
            del request
            obj = await ax_cache.cache.get('user_list')
            ret_str = 'READ cache == ' + \
                str(obj[0].username + ' - ' + os.environ['AX_VERSION'])
            return response.text(ret_str)

        @sanic_app.route('/api/console')
        async def console(request):  # pylint: disable=unused-variable
            """Cache Test function"""
            del request
            ax_pubsub.publisher.publish(
                aiopubsub.Key('console_log'), 'Hello world')
            
            return response.text('Pubsub sent')



        @sanic_app.route('/api/wine')
        async def wine(request):  # pylint: disable=unused-variable
            """Test function"""
            del request

            from pyquery import PyQuery as pq
            # from lxml import etree
            import urllib
            import re

            search_string = 'Chateau Pontet-Canet 2005'
            url_part = urllib.parse.quote_plus(search_string)

            search_url = f'http://www.wine.com/search/' + url_part + '/0?showOutOfStock=true'
            search_page = pq(url=search_url)
            wine_page_url = search_page(".prodItemInfo_link")[0].attrib['href']
            wine_page = pq(url='http://www.wine.com' + wine_page_url)

            initials = wine_page(
                ".pipProfessionalReviews_list .wineRatings_initials")
            ratings = wine_page(".wineRatings_rating")
            texts = wine_page(".pipSecContent_copy")

            dollars = pq(wine_page(".productPrice_price-saleWhole")).text()
            cents = pq(wine_page(".productPrice_price-saleFractional")).text()
            price = dollars + "." + cents

            label_small_img = wine_page(".pipThumb_img-0")[0].attrib['src']
            front_small_img = wine_page(".pipThumb_img-1")[0].attrib['src']
            back_small_img = wine_page(".pipThumb_img-2")[0].attrib['src']

            label_img_id = re.findall(r'/(\w+).jpg$', label_small_img)[0]
            front_img_id = re.findall(r'/(\w+).jpg$', front_small_img)[0]
            back_img_id = re.findall(r'/(\w+).jpg$', back_small_img)[0]

            name = pq(wine_page(".pipName")).text()
            vintage = re.findall(r'(\d+)$', name)[0]

            big_img_url = 'http://www.wine.com/product/images/w_767,c_fit,q_auto:good,fl_progressive/'

            ret_dict = {
                "search_string": search_string,
                "name": name,
                "vintage": vintage,
                "verietal": pq(wine_page(".prodItemInfo_varietal")).text(),
                "origin": pq(wine_page(".pipOrigin_link")).text(),
                "volume": pq(wine_page(".prodAlcoholVolume_text")[0]).text(),
                "alcohol": pq(wine_page(".prodAlcoholPercent_percent")[0]).text(),
                "price": price,
                "label_img": big_img_url + label_img_id + '.jpg',
                "front_img": big_img_url + front_img_id + '.jpg',
                "back_img": big_img_url + back_img_id + '.jpg'
            }
            for idx, pq_element in enumerate(initials):
                tag = pq(pq_element).text()
                ret_dict[tag] = {
                    "rating": pq(ratings[idx]).text(),
                    "text": pq(texts[idx]).text(),
                }
            test_str = json.dumps(ret_dict, sort_keys=True, indent=4)
            return response.html("<pre>" + test_str + "</pre>")

        @sanic_app.route('/api/excel')
        async def excel(request):  # pylint: disable=unused-variable
            """Test function"""
            del request
            import openpyxl
            from openpyxl.styles import colors
            from openpyxl.styles import Font
            from pyquery import PyQuery as pq
            # from lxml import etree
            import urllib
            import re

            wb = openpyxl.load_workbook(r'D:\Projects\wine_art_short.xlsx')
            ws = wb.active
            search_list = []

            # For each row of catalog -
            for row_index, row in enumerate(ws.iter_rows()):

                # Checks if row data is valid. If not ->
                # append error info to row.
                row_is_valid = True
                errors = []

                # Get row values
                name = str(row[0].value)
                vintage = str(row[1].value)
                volume = str(row[2].value)
                quantity = row[3].value
                price_string = str(row[4].value)

                volume_is_float = True
                try:
                    float(volume)
                except ValueError:
                    volume_is_float = False

                vintage_is_int = True
                vintage_clean = None
                try:
                    vintage_stipped = re.findall(r"\d+", vintage)[0]
                    vintage_clean = int(vintage_stipped)
                except Exception: # pylint: disable=broad-except
                    vintage_is_int = False

                # Work only with rows that contains quantity
                if not quantity or quantity == 0 or not name:
                    row_is_valid = False
                    errors.append(f'Quantity={quantity}')
                # If many years in cell -> we have a gift
                elif not vintage or not vintage_is_int:
                    row_is_valid = False
                    errors.append(
                        'Multiple years detected, gifts not supported')
                # If many years in cell -> we have a gift
                elif not volume or not volume_is_float:
                    row_is_valid = False
                    errors.append('Volume is wrong, gifts not supported')
                elif 'OWC-' in name:
                    row_is_valid = False
                    errors.append('OWC gifts not supported')

                if not row_is_valid:
                    msg = " ".join(errors)
                    row[6].value = msg
                    row[6].font = Font(color=colors.RED)
                else:
                    search_string = name + ' ' + str(vintage_clean)
                    stock_item = {
                        "rowIndex": row_index,
                        "searchString": search_string,
                        "vintage": str(vintage_clean),
                        "volume": volume,
                        "quantity": quantity,
                        "priceString": price_string,
                    }
                    search_list.append(stock_item)

            items_total = len(search_list)
            print(f"Found {items_total} stock items to search. Starting search")

            for idx, stock_item in enumerate(search_list):
                print(f"Parsing {idx}/{items_total}")
                url_part = urllib.parse.quote_plus(stock_item["searchString"])
                search_url = f'http://www.wine.com/search/' + url_part + '/0?showOutOfStock=true'
                search_page = pq(url=search_url)
                wine_page_url = search_page(".prodItemInfo_link")[0].attrib['href']
                wine_name = pq(search_page(".prodItemInfo_name")[0]).text()
                wine_vintage = re.findall(r'(\d+)$', wine_name)[0]

                search_is_valid = True
                # If nothing found - error
                if not wine_name:
                    search_is_valid = False
                    errors.append('No wine found')

                # Years do not match
                if str(wine_vintage) != str(stock_item['vintage']):
                    search_is_valid = False
                    errors.append('Years do not match')

                if not search_is_valid:
                    stock_item["error"] = " ".join(errors)
                    print(stock_item["error"])
                else:
                    stock_item["wineUrl"] = wine_page_url
                    stock_item["wineName"] = wine_name
                    print(f"{wine_name} OK")

            # Save results to excel -
            print("grabbing done, saving to excel")
            for row_index, row in enumerate(ws.iter_rows()):
                for stock_item in search_list:
                    if stock_item["rowIndex"] == row_index:
                        if "error" in stock_item:
                            row[6].value = stock_item["error"]
                            row[6].font = Font(color=colors.RED)
                            row[7].value = stock_item["searchString"]
                        else:
                            row[7].value = stock_item["searchString"]
                            row[8].value = stock_item["wineName"]
                            row[9].value = stock_item["wineUrl"]

                            if stock_item["searchString"] == stock_item["wineName"]:
                                row[7].font = Font(color=colors.GREEN)
                                row[8].font = Font(color=colors.GREEN)


            wb.save(r'D:\Projects\parse_result.xlsx')
            print('ALL DONE EXCEL SAVED')

            # Check if stock exists with this searchString
            # if exists -> get wine guid and skip Grab
            # else ->
                # Grab search page. Check if wine is found. Get wine id. If not -> append error info to row.
                # Check if wine id already grabbed, get wine data
                # Do create action on Wine
                    # Grab wine page -> get wine data
            # If Wine guid is set - create stock row -> write excel with wine data. Name, verietal, origin, volume
            # Create new excel file and write it to current loader row.

            test_str = json.dumps(search_list, sort_keys=True, indent=4)
            return response.html("<pre>" + test_str + "</pre>")

        @sanic_app.route('/api/graph')
        async def graph(request):  # pylint: disable=unused-variable
            """Test function"""
            del request

            job = ax_scheduler.scheduler.add_job(do_action, id='do_action_job')


            return response.html("<pre> Started" + job.name + "</pre>")

    except Exception:
        logger.exception('Error initiating routes.')
        raise
